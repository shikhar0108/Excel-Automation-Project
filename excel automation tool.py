import openpyxl as xl #For working with excel
from openpyxl.chart import BarChart,Reference


wb=xl.load_workbook("mall rate.xl.xlsx")
sheet=wb.active
for row in range(2,sheet.max_row+1):
    price=sheet[f"B{row}"].value
    dis_price = price*0.8
    sheet[f"C{row}"]= dis_price
print("applied successfully")

# For Bar Chart

chart=BarChart()
chart.title="Discount Price"
chart.x_axis.title="ITEM"
chart.y_axis.title="PRICE"

data = Reference (sheet,
                 min_col=2,
                 max_col=3,
                 min_row=1,
                 max_row=sheet.max_row)
category=Reference(sheet,
         min_col=1,
         min_row=2,
         max_row=sheet.max_row)


chart.add_data(data)
chart.set_categories(category)
sheet.add_chart(chart,"E2")
wb.save("mall rate.xl.xlsx")
